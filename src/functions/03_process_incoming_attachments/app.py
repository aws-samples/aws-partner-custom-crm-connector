from datetime import datetime
import json
import os
from io import BytesIO
import boto3
from aws_lambda_powertools import Logger, Tracer
from aws_lambda_powertools.logging.correlation_paths import S3_OBJECT_LAMBDA
from openpyxl import load_workbook

logger = Logger()
tracer = Tracer()
s3 = boto3.resource('s3')


@logger.inject_lambda_context(correlation_id_path=S3_OBJECT_LAMBDA)
def lambda_handler(event, context):
    try:
        print(event)
        bucket = event["S3"]["Bucket"]
        key = event["S3"]["Key"]
        entities = parse_entities(bucket, key)
        result = []
        for entity in entities:

            if "/opportunity/" in key:
                uniqueid=entity["partnerCrmUniqueIdentifier"]
                result.append(build_ace_json(
                    opportunity=entity, name=uniqueid))
            if "/lead/" in key:
                #uniqueid=tbd
                result.append(build_ace_json(lead=entity))
        print(result)
    finally:
        delete_object_from_s3(bucket, key)
    return result

@tracer.capture_method
def build_ace_json(lead: dict = None, opportunity: dict = None, name: str = None):
    __doc__ = "Build Skeleton for ACE JSON for one given lead or opportunity"

    if lead and opportunity:
        raise Exception("You can't pass both a lead and an opportunity")

    event = {
        "version": "1",
        "spmsId": os.environ['PARTNER_ID'],
        "name": name
    }
    if lead:
        event["leads"] = [lead]
    elif opportunity:
        event['opportunities'] = [opportunity]
    json_event = json.dumps(event, cls=DateTimeEncoder)

    return json.loads(json_event)

@tracer.capture_method
def delete_object_from_s3(s3bucket, s3key):
    s3_client = boto3.client('s3')

    print("Purge file")

    response = s3_client.delete_object(
        Bucket=s3bucket,
        Key=s3key
    )
    print(response)
    return response

@tracer.capture_method
def parse_entities(bucket, key) -> list:
    s3_object = s3.Object(bucket, key)
    print("Reading Excel")
    print("Bucket:" + bucket)
    print("Key:" + key)
    content = s3_object.get()['Body'].read()
    print("Trying to process tab 'Target'")
    ws = load_workbook(BytesIO(content), data_only=True)[
        'Target']
    headers = {}
    for row in ws.iter_rows(max_row=1):
        for column in row:
            headers[column.column_letter] = column.value
    result = []
    for row in ws.iter_rows(min_row=4):
        input = {}
        for column in row:
            key_name = headers[column.column_letter]
            if column.value:
                input[key_name] = column.value
        if len(input) > 0:
            result.append(input)
    return result


class DateTimeEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, datetime):
            return o.strftime("%Y-%m-%d")

        return json.JSONEncoder.default(self, o)