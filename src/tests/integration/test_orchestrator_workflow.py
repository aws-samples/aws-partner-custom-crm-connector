from datetime import datetime
import json
import urllib.request
from openpyxl import load_workbook
import email
import boto3
import pytest
from mypy_boto3_s3.service_resource import Bucket
from src.tests.config import Config
import time

sfn = boto3.client('stepfunctions')
STEPFUNCTION_ARN: str
STEPFUNCTION_ARN_ORCHESTRATION: str
bucket: Bucket
conf: Config
UPLOAD_PATH_EMAIL = "incoming-email"
UPLOAD_PATH_ATTACHEMENT = "incoming-attachments"


def setup_module(module):
    global conf
    global bucket
    conf = Config()
    if conf.stage == "prod":
        raise Exception("Invalid Stage for Integration Test")
    bucket = boto3.resource('s3').Bucket(conf.bucket_name)


@pytest.fixture
def email_msg(execution_name):
    msg = email.message_from_string(
        f"""From: "Lastname, Firstname" <{conf.email}>
To: "inbox@apn-ace-integration-framework-000000.awsapps.com"
	<inbox@apn-ace-integration-framework-000000.awsapps.com>
Subject: Trigger Test Execution
Date: Thu, 19 Jan 2023 07:56:32 +0000
Content-Type: multipart/mixed;
	boundary="somefilecontent_boundry"
MIME-Version: 1.0



--somefilecontent_boundry
Content-Type: application/octet-stream; name="{execution_name}"
Content-Description: {execution_name}
Content-Disposition: attachment; filename="{execution_name}";
	creation-date="Thu, 19 Jan 2023 07:56:32 GMT";
Content-Transfer-Encoding: base64


--somefilecontent_boundry--""")
    return msg


def test_upload_triggers_workflow(email_msg, execution_name):
    trigger = bucket.Object(
        f"{UPLOAD_PATH_EMAIL}/{execution_name}").put(Body=str(email_msg))
    time.sleep(3)
    result = sfn.list_executions(
        stateMachineArn=conf.step_function_orchestration_arn, statusFilter='RUNNING')

    executions = {}
    for execution in result['executions']:
        if execution_name in execution['name']:
            sfn.stop_execution(executionArn=execution['executionArn'])
            break
    else:
        raise AssertionError("No Running Execution Found")


@pytest.fixture(scope="module")
def workflow_results():
    str_date = datetime.now().strftime("%Y%d%m%H%M%S")
    test_name = f"integration-test-{str_date}"
    s3_key = f"{UPLOAD_PATH_ATTACHEMENT}/opportunity/{test_name}.xlsx"

    bucket.upload_file("ace_import_tmpl.xlsx", Key=s3_key)

    # Execute Workflow
    execution_arn = sfn.start_execution(
        stateMachineArn=conf.step_function_orchestration_arn,
        input=json.dumps({"S3": {"Bucket":  bucket.name, "Key": s3_key}}),
        name=test_name)['executionArn']
    manually_approved_given = False
    try:
        for i in range(30):
            time.sleep(2)
            result = sfn.describe_execution(executionArn=execution_arn)
            execution_history = sfn.get_execution_history(
                executionArn=execution_arn, reverseOrder=True)

            if not manually_approved_given and 'invoke.waitForTaskToken' == execution_history['events'][0].get('taskSubmittedEventDetails', {}).get('resource'):
                task_token = json.loads(execution_history['events'][2]['taskScheduledEventDetails']['parameters'])[
                    'Payload']['ExecutionContext']['Task']['Token']
                trigger_approval(task_token=task_token)
                manually_approved_given = True
            if result['status'] != "RUNNING":
                return {
                    "arn": execution_arn,
                    "status": result['status'],
                    "history": execution_history
                }
        else:
            raise TimeoutError("StepFunction still running")
    except Exception as ex:
        sfn.stop_execution(executionArn=execution_arn)
        raise


def data():
    ws = load_workbook("ace_import_tmpl.xlsx", data_only=True)['Target']
    tests_data = []
    for row in ws.iter_rows():
        if row[0].row > 3 and row[0].value:
            name = row[43].value
            description = row[2].value
            tests_data.append((name, "true" in description))

    return tests_data


@ pytest.mark.parametrize("execution_name,expected", data())
def test_opportunity_(execution_name, expected, workflow_results):
    result = json.loads(
        workflow_results['history']['events'][0]['executionSucceededEventDetails']['output'])

    assert expected == result[execution_name]['isSuccess']


def trigger_approval(task_token):
    approval_url = f"{conf.api_gateway_url}/approve?taskToken={urllib.parse.quote(task_token)}"
    contents = urllib.request.urlopen(approval_url).read()
    return contents
